# -*- coding: utf-8 -*-
#
# Copyright © 2012 - 2018 Michal Čihař <michal@cihar.com>
#
# This file is part of Weblate <https://weblate.org/>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
"""External file format specific behavior."""
from __future__ import unicode_literals

import csv
import os
import traceback
import re
from io import BytesIO
from zipfile import BadZipfile

import six
from django.utils.translation import ugettext_lazy as _

from openpyxl import Workbook, load_workbook

from weblate.trans.util import add_configuration_error

FILE_FORMATS = {}
FILE_DETECT = []


def register_external_fileformat(fileformat):
    """Register external fileformat in dictionary."""
    try:
        FILE_FORMATS[fileformat.format_id] = fileformat
        for autoload in fileformat.autoload:
            FILE_DETECT.append((autoload, fileformat))
    except (AttributeError, ImportError):
        add_configuration_error(
            'External file format: {0}'.format(fileformat.format_id),
            traceback.format_exc()
        )
    return fileformat


def detect_filename(filename):
    """Filename based format autodetection"""
    name = os.path.basename(filename)
    for autoload, storeclass in FILE_DETECT:
        if not isinstance(autoload, tuple) and name.endswith(autoload):
            return storeclass
        elif (name.startswith(autoload[0]) and
              name.endswith(autoload[1])):
            return storeclass
    return None


class ExternalFileFormat(object):
    name = ''
    format_id = ''
    autoload = ()

    def add_units(self, language, subproject, units):
        """add units to this translation file"""
        raise NotImplementedError()

    @staticmethod
    def convert_to_internal(filename, content):
        """convert the given content to an internal format"""
        raise NotImplementedError()


@register_external_fileformat
class XlsxFormat(ExternalFileFormat):
    name = _('Excel workbook')
    format_id = 'xlsx'
    autoload = ('.xlsx')
    workbook = Workbook()

    def add_units(self, language, subproject, units):
        worksheet = self.workbook.active

        if subproject is not None:
            worksheet.title = "{0}-{1}-{2}".format(
                subproject.project.slug,
                subproject.slug,
                language.code
            )
        else:
            worksheet.title = "{0}".format(
                language.code
            )

        # write headers
        worksheet.cell(
            column=1,
            row=1,
            value="source"
        )
        worksheet.cell(
            column=2,
            row=1,
            value="target"
        )
        worksheet.cell(
            column=3,
            row=1,
            value="context"
        )

        row = 2

        # use the same relugar expression as in openpyxl.cell
        # to remove illegal characters
        illegal_characters_re = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        for unit in units.iterator():
            # Write the translation data to the worksheet.
            # To suppress openpyxl to export values as formulas, we
            # set the cell value explicitly via .set_explicit_value.

            # first column is the source
            worksheet.cell(
                column=1,
                row=row,
            ).set_explicit_value(
                "{0}".format(illegal_characters_re.sub(r'', unit.source)),
                data_type="s"
            )

            # second column is the target
            worksheet.cell(
                column=2,
                row=row,
            ).set_explicit_value(
                "{0}".format(illegal_characters_re.sub(r'', unit.target)),
                data_type="s"
            )

            # third column is the context
            worksheet.cell(
                column=3,
                row=row,
            ).set_explicit_value(
                "{0}".format(illegal_characters_re.sub(r'', unit.context)),
                data_type="s"
            )

            row += 1

    def create_xlsx(self):
        output = BytesIO()
        self.workbook.save(output)
        return output.getvalue()

    @staticmethod
    def convert_to_internal(filename, content):
        # try to load the given file via openpyxl
        # catch at least the BadZipFile exception if an unsupported file has been given
        try:
            workbook = load_workbook(filename=BytesIO(content))
            worksheet = workbook.active
        except BadZipfile:
            return None, None

        output = BytesIO()
        writer = csv.writer(output)

        for row in worksheet.rows:
            writer.writerow([XlsxFormat.encode(cell.value) for cell in row])

        name = os.path.basename(filename) + ".csv"

        # return the new csv as bytes
        return name, output.getvalue()

    @staticmethod
    def encode(value):
        if value is None:
            return value
        if six.PY2:
            return value.encode("utf-8")
        return value

# Copyright © Michal Čihař <michal@weblate.org>
#
# SPDX-License-Identifier: GPL-3.0-or-later

"""External file format specific behavior."""

from __future__ import annotations

import os
from io import BytesIO, StringIO
from typing import TYPE_CHECKING
from zipfile import BadZipFile

from django.utils.translation import gettext_lazy
from translate.storage.csvl10n import csv

from weblate.formats.helpers import CONTROLCHARS_TRANS, NamedBytesIO
from weblate.formats.ttkit import CSVUtf8Format

if TYPE_CHECKING:
    from collections.abc import Callable

CSV_DIALECT = "unix"


class XlsxFormat(CSVUtf8Format):
    name = gettext_lazy("Excel Open XML")
    format_id = "xlsx"
    autoload = ("*.xlsx",)

    def write_cell(self, worksheet, column: int, row: int, value: str):
        from openpyxl.cell.cell import TYPE_STRING

        cell = worksheet.cell(column=column, row=row)
        cell.value = value
        # Set the data_type after value to override function auto-detection
        cell.data_type = TYPE_STRING
        return cell

    def get_title(self, fallback: str = "Weblate"):
        from openpyxl.workbook.child import INVALID_TITLE_REGEX

        title = self.store.targetlanguage
        if title is None:
            return fallback
        # Remove possible invalid characters
        title = INVALID_TITLE_REGEX.sub(title, "").strip()
        if not title:
            return fallback
        return title

    def save_content(self, handle) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.get_title()
        fieldnames = self.store.fieldnames

        # write headers
        for column, field in enumerate(fieldnames):
            self.write_cell(worksheet, column + 1, 1, field)

        for row, unit in enumerate(self.store.units):
            data = unit.todict()

            for column, field in enumerate(fieldnames):
                self.write_cell(
                    worksheet,
                    column + 1,
                    row + 2,
                    data[field].translate(CONTROLCHARS_TRANS),
                )

        workbook.save(handle)

    @staticmethod
    def serialize(store):
        # store is CSV (csvfile) here
        output = BytesIO()
        XlsxFormat(store).save_content(output)
        return output.getvalue()

    def parse_store(self, storefile):
        from openpyxl import load_workbook

        # try to load the given file via openpyxl
        # catch at least the BadZipFile exception if an unsupported
        # file has been given
        try:
            workbook = load_workbook(filename=storefile)
            worksheet = workbook.active
        except BadZipFile:
            return None, None

        output = StringIO()

        writer = csv.writer(output, dialect=CSV_DIALECT)

        # value can be None or blank stringfor cells having formatting only,
        # we need to ignore such columns as that would be treated like "" fields
        # later in the translate-toolkit
        fields = [cell.value for cell in next(worksheet.rows) if cell.value]
        for row in worksheet.rows:
            values = [cell.value for cell in row]
            values = values[: len(fields)]
            # Skip formatting only cells
            if not any(values):
                continue
            writer.writerow(values)

        if isinstance(storefile, str):
            name = os.path.basename(storefile) + ".csv"
        else:
            name = os.path.basename(storefile.name) + ".csv"

        # return the new csv as bytes
        content = output.getvalue().encode("utf-8")

        # Load the file as CSV
        return super().parse_store(NamedBytesIO(name, content), dialect=CSV_DIALECT)

    @staticmethod
    def mimetype() -> str:
        """Return most common mime type for format."""
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @staticmethod
    def extension() -> str:
        """Return most common file extension for format."""
        return "xlsx"

    @classmethod
    def create_new_file(
        cls,
        filename: str,
        language: str,
        base: str,
        callback: Callable | None = None,
    ) -> None:
        """Handle creation of new translation file."""
        if not base:
            msg = "Not supported"
            raise ValueError(msg)
        # Parse file
        store = cls(base)
        if callback:
            callback(store)
        store.untranslate_store(language)
        with open(filename, "wb") as handle:
            XlsxFormat(store.store).save_content(handle)

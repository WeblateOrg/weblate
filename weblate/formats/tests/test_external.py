# Copyright © Michal Čihař <michal@weblate.org>
#
# SPDX-License-Identifier: GPL-3.0-or-later

"""File format specific behavior."""

import tempfile
from io import BytesIO

from openpyxl import load_workbook

from weblate.formats.external import XlsxFormat
from weblate.formats.tests.test_formats import BaseFormatTest
from weblate.trans.tests.utils import get_test_file

XLSX_FILE = get_test_file("cs-mono.xlsx")
JAPANESE_FILE = get_test_file("ja.xlsx")
FRENCH_FILE = get_test_file("fr.xlsx")


class XlsxFormatTest(BaseFormatTest):
    format_class = XlsxFormat
    FILE = XLSX_FILE
    MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    EXT = "xlsx"
    COUNT = 4
    MASK = "*/translations.xlsx"
    EXPECTED_PATH = "cs_CZ/translations.xlsx"
    FIND = "HELLO"
    FIND_MATCH = "Hello, world!\r\n"
    MATCH = b"PK"
    NEW_UNIT_MATCH = b"PK"
    BASE = XLSX_FILE
    EXPECTED_FLAGS = ""

    def assert_same(self, newdata, testdata) -> None:
        newworkbook = load_workbook(BytesIO(newdata))
        testworkbook = load_workbook(BytesIO(testdata))
        self.assertEqual(len(newworkbook.worksheets), len(testworkbook.worksheets))
        self.assertIsNotNone(newworkbook.active)
        self.assertIsNotNone(testworkbook.active)
        self.assertEqual(
            list(newworkbook.active.values),
            list(testworkbook.active.values),
        )

    def test_japanese(self) -> None:
        storage = self.format_class(JAPANESE_FILE)
        self.assertEqual(len(storage.all_units), 1)
        self.assertEqual(storage.all_units[0].target, "秒")
        with tempfile.NamedTemporaryFile(suffix="xlsx") as temp_file:
            storage.save_atomic(temp_file.name, storage.save_content)

    def test_fr(self) -> None:
        storage = self.format_class(FRENCH_FILE)
        self.assertEqual(len(storage.all_units), 4)
        self.assertEqual(storage.all_units[0].target, "Traitement A")
        with tempfile.NamedTemporaryFile(suffix="xlsx") as temp_file:
            storage.save_atomic(temp_file.name, storage.save_content)
